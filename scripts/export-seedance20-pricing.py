"""Export Seedance 2.0 pricing tables to Excel."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parents[1] / "Seedance2.0定价表.xlsx"
MARKUP = 1.9  # 上浮 90%
DURATIONS = list(range(4, 16))

STANDARD_RATES = {
    "480P": 100,
    "720P": 120,
    "1080P": 300,
    "4K": 600,
}

FAST_RATES = {
    "480P": 80.6,
    "720P": 96.6,
}


def credits(rate: float, seconds: int, *, round_result: bool = False) -> int:
    value = rate * seconds
    return int(round(value)) if round_result else int(value)


def yuan(credits_value: int) -> float:
    return round(credits_value / 100, 2)


def style_header_row(ws, row: int, col_count: int) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, min_width: int = 10, max_width: int = 18) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


def build_standard_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "标准版 seedance-2.0"

    ws["A1"] = "Seedance 2.0 标准版（seedance-2.0）"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:M1")

    ws["A2"] = "单价：480P=100积分/秒，720P=120，1080P=300，4K=600；100积分=1元；上浮90%=原价×1.9"
    ws.merge_cells("A2:M2")

    headers = ["秒数"]
    for res in STANDARD_RATES:
        headers.extend([f"{res} 原积分", f"{res} 上浮积分", f"{res} 上浮¥"])
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    style_header_row(ws, header_row, len(headers))

    for sec in DURATIONS:
        row = [sec]
        for rate in STANDARD_RATES.values():
            base = credits(rate, sec)
            marked = int(round(base * MARKUP))
            row.extend([base, marked, yuan(marked)])
        ws.append(row)

    auto_width(ws)


def build_fast_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("快速版 seedance-2.0-fast")

    ws["A1"] = "Seedance 2.0 快速版（seedance-2.0-fast）"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")

    ws["A2"] = "单价：480P=80.6积分/秒，720P=96.6；积分四舍五入取整；上浮90%=原价×1.9"
    ws.merge_cells("A2:G2")

    headers = ["秒数"]
    for res in FAST_RATES:
        headers.extend([f"{res} 原积分", f"{res} 上浮积分", f"{res} 上浮¥"])
    ws.append([])
    ws.append(headers)
    style_header_row(ws, ws.max_row, len(headers))

    for sec in DURATIONS:
        row = [sec]
        for rate in FAST_RATES.values():
            base = credits(rate, sec, round_result=True)
            marked = int(round(base * MARKUP))
            row.extend([base, marked, yuan(marked)])
        ws.append(row)

    auto_width(ws)


def main() -> None:
    wb = Workbook()
    build_standard_sheet(wb)
    build_fast_sheet(wb)
    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    main()
